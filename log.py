import os
import openpyxl


class XLLogSaver(object):
    def __init__(self, xl_folder_name, xl_file_name, tabs):
        super().__init__()

        self.xl_folder_name = xl_folder_name
        self.xl_file_name = xl_file_name
        self.tabs = tabs

        self.wb = None
        self.ws = None
        self.init_xl_saver()

    def init_xl_saver(self):

        os.makedirs(self.xl_folder_name, exist_ok=True)
        # self.check_duplication()

        self.wb = openpyxl.Workbook()
        self.wb.save(os.path.join(self.xl_folder_name, self.xl_file_name + '.xlsx'))
        self.ws = self.wb.active

        self.insert_header()

    def check_duplication(self):
        assert not os.path.isfile(os.path.join(self.xl_folder_name, self.xl_file_name + '.xlsx')), \
            'please check not to overload file.'

    def insert_header(self):
        sub = self.tabs
        for kwd, j in zip(sub, list(range(1, len(sub) + 1))):
            self.ws.cell(row=j, column=1).value = kwd
        self.wb.save(os.path.join(self.xl_folder_name, self.xl_file_name + '.xlsx'))

    def insert_each_epoch(self, contents):
        assert len(contents) == len(self.tabs), 'must have the same length between tabs and contents'
        epoch = int(contents[0])
        for kwd, j in zip(contents, list(range(1, len(contents) + 1))):
            self.ws.cell(row=j, column=epoch + 2).value = kwd
        self.wb.save(os.path.join(self.xl_folder_name, self.xl_file_name + '.xlsx'))